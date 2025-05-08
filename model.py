import pandas as pd
import numpy as np
import statsmodels.formula.api as smf
import statsmodels.api as sm
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

# Create directory for results
os.makedirs('results', exist_ok=True)

# Load data
print("Loading data...")
df = pd.read_csv('data/Combined_Joined_Data.csv')

# Model configuration - EDIT THIS SECTION to customize your models
model_config = {
    # Main experimental effects to always include (required)
    'main_effects': ['IsGreen', 'RunGroupType', 'RunOrder_num'],
    
    # Interaction terms to include (set to empty list [] to exclude all interactions)
    'interactions': [
        # Specify as tuples of variables that should interact
        #('IsGreen', 'RunGroupType'),  # Comment or delete to exclude IsGreen:RunGroupType interaction
    ],
    
    # Covariates to include (set to empty list [] to exclude all covariates)
    'covariates': [
        #'Age', 
        #'Gender_num', 
        #'VO2max', 
        #'AvgSkinTemp', 
        # 'Presence_Avg',  # Uncomment to include Presence_Avg
    ],
    
    # Use random intercept for participant
    'random_intercept': 'ParticipantID'
}

# Print columns for reference
print("\nAvailable columns:")
columns_list = df.columns.tolist()
print(columns_list)

# Define column mappings with correct names from the data
column_mappings = {
    # Basic info columns
    'experiment ID (participant number_condition)': 'ExperimentID',
    'participant number': 'ParticipantID',
    'condition': 'Condition_num',
    'run order': 'RunOrder_num',
    'Past exepriences with XR': 'MRExperience_num',
    'Gender': 'Gender_num',
    'Age': 'Age',
    'VO2max': 'VO2max',
    'avg_close_to_skin_temp': 'AvgSkinTemp',
    'min_heart_rate': 'MinHR',
    'Average heart rate (bpm)': 'AvgHR',
    'HR max': 'HR_max',
    'Recalculated_Speed_km_per_h': 'Speed_kmh',
    'Duration_seconds': 'Duration_s',
    
    # Key measure columns with correct names from the data
    'RPE': 'RPE',
    'Satisfaction': 'Satisfaction',
    'Enjoyment': 'Enjoyment',
    'Challenging': 'Challenging',
    'Concentration': 'Concentration',
    'Feeling of presence': 'Presence', 
    'Ability to move freely': 'MoveFreely',
    'Fatigue': 'Fatigue',
    'Exhaustion': 'Exhaustion',
    'Sluggish': 'Sluggish',
    'Feeling light weighted': 'LightWeighted',
    'Relaxation and losing oneself in the environment': 'Relaxation',
    'Curiosity and fascination': 'Curiosity',
    'Order in space': 'OrderInSpace',
    'Easy navigation': 'EasyNavigation',
    'Comfortable environment': 'ComfortableEnv',
    'Tense': 'Tense',
    'Anxious': 'Anxious',
    'Healthy': 'Healthy',
    'Angry': 'Angry',
    'Irritated': 'Irritated',
    'Immersed': 'Immersed',
    'Physically present': 'PhysicallyPresent',
    'Movement around and use of objects': 'UseObjects',
    'Ability to do things': 'AbilityToDo',
    'Identification': 'Identification',
    'Greenes of the environment': 'PerceivedGreenness',
    'Sense of belonging': 'Belonging',
    
    # Add missing environmental perception variables with proper mappings
    'Interesting things to observe': 'InterestingObserve',
    'Monotonous': 'Monotonous',
    'Not well-maintained': 'NotMaintained'
}

# Rename Columns
rename_dict = {k: v for k, v in column_mappings.items() if k in df.columns}
df = df.rename(columns=rename_dict)
# Print renamed columns for reference
print("\nRenamed columns:")
print(df.columns.tolist())

# Convert numeric columns to float
numeric_columns = [col for col in df.columns if col in set(column_mappings.values())]
print(f"\nProcessing {len(numeric_columns)} numeric columns.")
for col in numeric_columns:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')


df['PercentMaxHR'] = df.apply(lambda row: (row['AvgHR'] / row['HR_max'] * 100), axis=1)
print("\nCalculated PercentMaxHR")

# Ensure Condition_num is numeric
df['Condition_num'] = pd.to_numeric(df['Condition_num'], errors='coerce')

# Determine if a participant was in tree group (2) or shrub group (1)
participant_groups = {}

# For each participant, check their condition assignments
for participant_id in df['ParticipantID'].unique():
    participant_data = df[df['ParticipantID'] == participant_id]
    
    # Check if they have tree runs (condition 2)
    has_tree_runs = (participant_data['Condition_num'] == 2).any()
    
    # Check if they have shrub runs (condition 1)
    has_shrub_runs = (participant_data['Condition_num'] == 1).any()
    
    # Assign group (2 for tree group, 1 for shrub group)
    if has_tree_runs and not has_shrub_runs:
        participant_groups[participant_id] = 2  # Tree group
    elif has_shrub_runs and not has_tree_runs:
        participant_groups[participant_id] = 1  # Shrub group

# Create RunGroupType column (Tree=2, Shrub=1)
df['RunGroupType'] = df['ParticipantID'].map(participant_groups)

# Create green condition indicator (1=green, 0=control)
df['IsGreen'] = (df['Condition_num'] > 0.5).astype(int)

# Print condition counts
print("\nCondition counts:")
print(f"Condition_num:\n{df['Condition_num'].value_counts().sort_index()}")
print(f"\nIsGreen (1=green, 0=control):\n{df['IsGreen'].value_counts()}")
print(f"\nRunGroupType (2=tree group, 1=shrub group):\n{df['RunGroupType'].value_counts()}")

# Calculate composite variables
print("\nCalculating composite variables...")
# Presence composite
presence_items = ['Immersed', 'PhysicallyPresent', 'UseObjects', 'AbilityToDo']
df['Presence_Avg'] = df[presence_items].mean(axis=1)
# Perceived Restorativeness Scale
prs_items = ['Relaxation', 'Curiosity', 'OrderInSpace', 'EasyNavigation', 'ComfortableEnv']
df['PRS_Avg'] = df[prs_items].mean(axis=1)
# Place Attachment
pa_items = ['Identification', 'Belonging']
df['PlaceAttach_Avg'] = df[pa_items].mean(axis=1)
# Positive Affect
positive_affect_items = ['Enjoyment', 'Satisfaction', 'Healthy', 'LightWeighted']
df['PositiveAffect_Avg'] = df[positive_affect_items].mean(axis=1)
# Negative Affect
negative_affect_items = ['Fatigue', 'Tense', 'Anxious', 'Angry', 'Irritated', 'Sluggish']
df['NegativeAffect_Avg'] = df[negative_affect_items].mean(axis=1)

# Define all dependent variables - including BOTH composite scales AND individual items
dep_vars = [
    # Composite scales
    'PRS_Avg', 'PlaceAttach_Avg', 'PositiveAffect_Avg', 'NegativeAffect_Avg', 'Presence_Avg',
    
    # Presence scale items (already included in original list)
    'Presence', 'Immersed', 'PhysicallyPresent', 'UseObjects', 'AbilityToDo',
    
    # PRS scale items
    'Relaxation', 'Curiosity', 'OrderInSpace', 'EasyNavigation', 'ComfortableEnv',
    
    # Place Attachment items
    'Identification', 'Belonging',
    
    # Positive Affect items
    'Enjoyment', 'Satisfaction', 'Healthy', 'LightWeighted',
    
    # Negative Affect items
    'Fatigue', 'Tense', 'Anxious', 'Angry', 'Irritated', 'Sluggish',
    
    # Other individual items
    'Duration_s', 'RPE', 'Concentration', 'MoveFreely', 'Exhaustion', 
    'Challenging', 'PerceivedGreenness', 'MinHR', 'Speed_kmh', 'AvgHR', 'PercentMaxHR',
    
    # Environmental perception items
    'InterestingObserve', 'Monotonous', 'NotMaintained'
]

# Check for duplicates
dep_vars = list(dict.fromkeys(dep_vars))
print(f"\nAnalyzing {len(dep_vars)} dependent variables: {dep_vars}")

# Define covariates
covariates = ['RunOrder_num', 'Age', 'Gender_num', 'VO2max', 'AvgSkinTemp', 'Presence_Avg']
print(f"Using {len(covariates)} covariates: {covariates}")

# Storage for results
all_results = {}

def get_significance_symbol(p_value):
    """Return significance symbols based on p-value"""
    if pd.isna(p_value):
        return ''
    if p_value < 0.001:
        return '***'
    elif p_value < 0.01:
        return '**'
    elif p_value < 0.05:
        return '*'
    elif p_value < 0.1:
        return '†'
    else:
        return ''

def run_mixed_model(data, dep_var):
    """Run a linear mixed effects model"""
    print(f"\nAnalyzing {dep_var}")
    
    # Prepare model data
    model_data = data.dropna(subset=[dep_var]).copy()
    
    # Skip model if insufficient data
    if len(model_data) < 10:
        print(f"  Skipping: Not enough data (N={len(model_data)})")
        return None
    
    # Ensure ParticipantID is a string for grouping
    model_data['ParticipantID'] = model_data['ParticipantID'].astype(str)
    
    # Start building the formula with main effects
    formula_parts = [f"{dep_var} ~"]
    
    # Add main effects
    valid_main_effects = [effect for effect in model_config['main_effects'] 
                         if effect in model_data.columns and effect != dep_var]
    if valid_main_effects:
        formula_parts.append(" + ".join(valid_main_effects))
    
    # Add interaction terms
    for var1, var2 in model_config['interactions']:
        if var1 in model_data.columns and var2 in model_data.columns and var1 != dep_var and var2 != dep_var:
            formula_parts.append(f"{var1}:{var2}")
    
    # Add covariates
    valid_covariates = [cov for cov in model_config['covariates'] 
                       if cov in model_data.columns and cov != dep_var
                       and model_data[cov].notna().sum() > 0.8 * len(model_data)]
    if valid_covariates:
        formula_parts.append(" + ".join(valid_covariates))
    
    # Join all parts with +
    formula = " + ".join(formula_parts)
    
    print(f"  Formula: {formula}")
    print(f"  N = {len(model_data)}, Participants = {model_data['ParticipantID'].nunique()}")
    
    try:
        # Run mixed model with random intercept for participant
        random_intercept = model_config['random_intercept']
        if random_intercept in model_data.columns:
            model = smf.mixedlm(formula, model_data, groups=model_data[random_intercept])
            fit = model.fit(reml=True)
            
            # Extract key parameters
            results = {
                'n': len(model_data),
                'participants': model_data['ParticipantID'].nunique(),
                'formula': formula,
                'fit': fit,
                'intercept': fit.params.get('Intercept', np.nan),
                'intercept_se': fit.bse.get('Intercept', np.nan),
                'intercept_p': fit.pvalues.get('Intercept', np.nan),
            }
            
            # Get coefficients for all variables
            for var in valid_main_effects + valid_covariates:
                results[f'{var}_effect'] = fit.params.get(var, np.nan)
                results[f'{var}_p'] = fit.pvalues.get(var, np.nan)
                results[f'{var}_se'] = fit.bse.get(var, np.nan)
                results[f'{var}_sig'] = get_significance_symbol(fit.pvalues.get(var, np.nan))
            
            # Get coefficients for interaction terms
            for var1, var2 in model_config['interactions']:
                interaction_term = f"{var1}:{var2}"
                if interaction_term in fit.params:
                    results[f'{var1}_{var2}_interaction'] = fit.params.get(interaction_term, np.nan)
                    results[f'{var1}_{var2}_interaction_p'] = fit.pvalues.get(interaction_term, np.nan)
                    results[f'{var1}_{var2}_interaction_se'] = fit.bse.get(interaction_term, np.nan)
                    results[f'{var1}_{var2}_interaction_sig'] = get_significance_symbol(fit.pvalues.get(interaction_term, np.nan))
            
            # Get random effects variance
            random_var = float(fit.cov_re.iloc[0, 0])
            residual_var = fit.scale
            icc = random_var / (random_var + residual_var)
            
            # Store random effects for each participant
            random_effects = fit.random_effects
            # Safely extract random effects values
            random_effects_values = []
            for participant, re_value in random_effects.items():
                # Extract the value (could be in different formats depending on statsmodels version)
                if hasattr(re_value, 'iloc') and re_value.size > 0:  # DataFrame or Series
                    value = float(re_value.iloc[0])
                elif hasattr(re_value, 'item'):  # numpy array
                    value = float(re_value.item())
                elif isinstance(re_value, (int, float)):
                    value = float(re_value)
                else:
                    # Try to convert to float if possible
                    try:
                        value = float(re_value)
                    except (TypeError, ValueError):
                        print(f"Warning: Could not convert random effect to float: {re_value}")
                        value = 0.0
                random_effects_values.append(value)
            
            # Calculate statistics on the extracted values
            random_effects_stats = {
                'mean': float(np.mean(random_effects_values)) if random_effects_values else 0.0,
                'sd': float(np.std(random_effects_values)) if random_effects_values else 0.0,
                'min': float(np.min(random_effects_values)) if random_effects_values else 0.0,
                'max': float(np.max(random_effects_values)) if random_effects_values else 0.0
            }
            
            results['random_effects'] = random_effects_stats
            results['participant_icc'] = icc
            
            # Calculate simplified R² values
            total_var = model_data[dep_var].var()
            residuals = model_data[dep_var] - fit.fittedvalues
            residual_var_actual = np.var(residuals)
            
            # Marginal R² - fixed effects only
            r2_marginal = 1 - (residual_var_actual / total_var)
            
            # Conditional R² - fixed + random effects 
            # Note: This is an approximation
            r2_conditional = 1 - (residual_var / total_var)
            
            # Make sure R² values are within bounds
            results['r2_marginal'] = max(0, min(1, r2_marginal))
            results['r2_conditional'] = max(0, min(1, r2_conditional))
            
            # Calculate AIC and BIC manually
            # Count parameters: fixed effects + 1 for random intercept variance + 1 for residual variance
            n_fixed_params = len(fit.params)
            k = n_fixed_params + 2  # +2 for random intercept variance and residual variance
            n = len(model_data)
            loglik = fit.llf
            results['aic'] = -2 * loglik + 2 * k
            results['bic'] = -2 * loglik + k * np.log(n)
            results['loglik'] = loglik
            
            # Print key results
            print(f"  Results:")
            print(f"    Intercept: {results['intercept']:.3f}")
            
            # Print main effects
            for var in valid_main_effects:
                effect_name = f"{var}_effect"
                p_name = f"{var}_p"
                sig_name = f"{var}_sig"
                if effect_name in results and p_name in results:
                    print(f"    {var} effect: {results[effect_name]:.3f} {results[sig_name]} (p={results[p_name]:.4f})")
            
            # Print interaction effects
            for var1, var2 in model_config['interactions']:
                interaction_name = f"{var1}_{var2}_interaction"
                p_name = f"{var1}_{var2}_interaction_p"
                sig_name = f"{var1}_{var2}_interaction_sig"
                if interaction_name in results and p_name in results:
                    print(f"    {var1}:{var2} interaction: {results[interaction_name]:.3f} {results[sig_name]} (p={results[p_name]:.4f})")
            
            # Print model fit statistics
            print(f"    Participant_ICC: {results['participant_icc']:.3f}, R² Marginal: {results['r2_marginal']:.3f}, R² Conditional: {results['r2_conditional']:.3f}")
            print(f"    AIC: {results['aic']:.2f}, BIC: {results['bic']:.2f}")
            
            return results
        else:
            print(f"  Error: Random intercept variable {random_intercept} not found in data")
            return None
    except Exception as e:
        print(f"  Error: {str(e)}")
        return None

def create_excel_report(results):
    """Create a clean, simple Excel report with results"""
    print("\nCreating Excel report...")
    
    if not results:
        print("No results to report")
        return
    
    try:
        excel_path = 'results/mixed_effects_model.xlsx'
        
        # Use context manager to ensure proper cleanup
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Create styles
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="E0EBF5", end_color="E0EBF5", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Extract all parameter names from results
            all_params = set()
            for res in results.values():
                for key in res.keys():
                    if key not in ['fit', 'random_effects', 'formula', 'n', 'participants', 'intercept', 'intercept_se', 'intercept_p']:
                        if key.endswith('_effect') or key.endswith('_interaction'):
                            all_params.add(key)
            
            # Create summary rows with all effects and p-values
            summary_rows = []
            for dv, res in results.items():
                # Start with basic info
                row = {
                    'Dependent Variable': dv,
                    'N': res['n'],
                    'Participants': res['participants'],
                    # Intercept and random effects
                    'Intercept': res['intercept'],
                    'Random_Effects_Mean': res['random_effects']['mean'],
                    'Random_Effects_SD': res['random_effects']['sd'],
                }
                
                # Add all parameters
                for param in all_params:
                    if param in res:
                        row[param] = res[param]
                        # Add p-value and significance if available
                        p_key = param.replace('_effect', '_p').replace('_interaction', '_interaction_p')
                        sig_key = param.replace('_effect', '_sig').replace('_interaction', '_interaction_sig')
                        if p_key in res:
                            row[p_key] = res[p_key]
                        if sig_key in res:
                            row[sig_key] = res[sig_key]
                
                # Add model fit metrics at the end
                row['R² Marginal'] = res['r2_marginal']
                row['R² Conditional'] = res['r2_conditional']
                row['Participant_ICC'] = res['participant_icc']
                row['AIC'] = res['aic']
                row['BIC'] = res['bic']
                
                summary_rows.append(row)
            
            # Sort by p-value of first effect (usually IsGreen)
            first_param = next(iter(all_params)) if all_params else None
            if first_param:
                p_key = first_param.replace('_effect', '_p').replace('_interaction', '_interaction_p')
                sig_key = first_param.replace('_effect', '_sig').replace('_interaction', '_interaction_sig')
                
                summary_df = pd.DataFrame(summary_rows)
                try:
                    if sig_key in summary_df.columns:
                        summary_df = summary_df.sort_values([sig_key, p_key], 
                                                key=lambda x: pd.Categorical(summary_df[sig_key], 
                                                                        categories=['***', '**', '*', '†', ''], 
                                                                        ordered=True),
                                                ascending=[False, True])
                except:
                    # Fallback if sort fails
                    pass
            else:
                summary_df = pd.DataFrame(summary_rows)
            
            # Write summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            ws = writer.sheets['Summary']
            
            # Format header row (no merged cells)
            try:
                for col in range(1, len(summary_df.columns) + 1):
                    # Format header cells
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    
                    # Set column width
                    ws.column_dimensions[get_column_letter(col)].width = 12
            except Exception as e:
                print(f"Warning: Error formatting header row: {str(e)}")
            
            # Special width for first column
            try:
                ws.column_dimensions['A'].width = 20
            except Exception as e:
                print(f"Warning: Error setting column A width: {str(e)}")
            
            # Add filter
            try:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(summary_df.columns))}1"
            except Exception as e:
                print(f"Warning: Error setting filter: {str(e)}")
            
            # Individual model sheets (compact)
            for dv, res in results.items():
                try:
                    # Create safe sheet name (max 31 chars)
                    sheet_name = dv[:30]
                    
                    # Create compact data for model sheet
                    model_data = [
                        ["Mixed Effects Model", dv, "", "", ""],
                        ["N:", f"{res['n']}", "Participants:", f"{res['participants']}", ""],
                        ["Formula:", f"{res['formula']}", "", "", ""],
                        ["", "", "", "", ""],
                        ["Model Components", "Value", "", "", ""],
                        ["Intercept", f"{res['intercept']:.4f}", "", "", ""],
                        ["Random Effects Mean", f"{res['random_effects']['mean']:.4f}", "", "", ""],
                        ["Random Effects SD", f"{res['random_effects']['sd']:.4f}", "", "", ""],
                        ["Min Random Effect", f"{res['random_effects']['min']:.4f}", "", "", ""],
                        ["Max Random Effect", f"{res['random_effects']['max']:.4f}", "", "", ""],
                        ["", "", "", "", ""],
                        ["Fixed Effects", "Estimate", "SE", "p-value", "Significance"],
                    ]
                    
                    # Add main effects
                    for var in model_config['main_effects']:
                        effect_key = f"{var}_effect"
                        se_key = f"{var}_se"
                        p_key = f"{var}_p"
                        sig_key = f"{var}_sig"
                        if effect_key in res:
                            model_data.append([
                                var, 
                                f"{res[effect_key]:.4f}", 
                                f"{res[se_key]:.4f}" if se_key in res else "",
                                f"{res[p_key]:.4f}" if p_key in res else "", 
                                res[sig_key] if sig_key in res else ""
                            ])
                    
                    # Add interaction terms
                    for var1, var2 in model_config['interactions']:
                        interaction_key = f"{var1}_{var2}_interaction"
                        se_key = f"{var1}_{var2}_interaction_se"
                        p_key = f"{var1}_{var2}_interaction_p"
                        sig_key = f"{var1}_{var2}_interaction_sig"
                        if interaction_key in res:
                            model_data.append([
                                f"{var1}:{var2}", 
                                f"{res[interaction_key]:.4f}", 
                                f"{res[se_key]:.4f}" if se_key in res else "",
                                f"{res[p_key]:.4f}" if p_key in res else "", 
                                res[sig_key] if sig_key in res else ""
                            ])
                    
                    # Add covariate effects
                    if model_config['covariates']:
                        model_data.append(["", "", "", "", ""])
                        model_data.append(["Covariates", "Estimate", "SE", "p-value", "Significance"])
                        
                        for cov in model_config['covariates']:
                            effect_key = f"{cov}_effect"
                            se_key = f"{cov}_se"
                            p_key = f"{cov}_p"
                            sig_key = f"{cov}_sig"
                            if effect_key in res:
                                model_data.append([
                                    cov, 
                                    f"{res[effect_key]:.4f}", 
                                    f"{res[se_key]:.4f}" if se_key in res else "",
                                    f"{res[p_key]:.4f}" if p_key in res else "", 
                                    res[sig_key] if sig_key in res else ""
                                ])
                    
                    # Add model fit stats
                    model_data.append(["", "", "", "", ""])
                    model_data.append(["Model Fit", "", "", "", ""])
                    model_data.append(["R² Marginal:", f"{res['r2_marginal']:.4f}", "R² Conditional:", f"{res['r2_conditional']:.4f}", ""])
                    model_data.append(["Participant_ICC:", f"{res['participant_icc']:.4f}", "Log-Likelihood:", f"{res['loglik']:.2f}", ""])
                    model_data.append(["AIC:", f"{res['aic']:.2f}", "BIC:", f"{res['bic']:.2f}", ""])
                    
                    # Convert to DataFrame and write to sheet
                    model_df = pd.DataFrame(model_data)
                    model_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    
                    # Format model sheet
                    ws = writer.sheets[sheet_name]
                    
                    # Format headers and section titles
                    for row_idx, row_data in enumerate(model_data, 1):
                        try:
                            if row_idx == 1:  # Title row
                                cell = ws.cell(row=row_idx, column=1)
                                cell.font = Font(bold=True, size=12)
                                
                                cell = ws.cell(row=row_idx, column=2)
                                cell.font = Font(bold=True, size=12)
                                
                            elif row_data[0] in ["Model Components", "Fixed Effects", "Model Fit", "Covariates"]:  # Section headers
                                cell = ws.cell(row=row_idx, column=1)
                                cell.font = header_font
                                cell.fill = header_fill
                                
                                # Format header row
                                for col_idx in range(1, 6):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    cell.font = header_font
                                    cell.fill = header_fill
                        except Exception as e:
                            print(f"Warning: Error formatting row {row_idx} in sheet {sheet_name}: {str(e)}")
                    
                    # Set column widths
                    try:
                        ws.column_dimensions['A'].width = 15
                        ws.column_dimensions['B'].width = 12
                        ws.column_dimensions['C'].width = 15
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 12
                    except Exception as e:
                        print(f"Warning: Error setting column widths in sheet {sheet_name}: {str(e)}")
                except Exception as e:
                    print(f"Warning: Error creating sheet for {dv}: {str(e)}")
        
        print(f"Excel report saved to {excel_path}")
    except Exception as e:
        print(f"Error creating Excel report: {str(e)}")

# Run models for all dependent variables
print("\nRunning mixed effects models for all dependent variables...")
for dv in dep_vars:
    results = run_mixed_model(df, dv)
    if results:
        all_results[dv] = results

# Generate the report
create_excel_report(all_results)

print("\nAnalysis complete!")